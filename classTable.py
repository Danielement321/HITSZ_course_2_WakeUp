import openpyxl
import openai
from tqdm import tqdm

BASE_URL = '' # Your agent url
API_KEY = 'sk-xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx' # Your API key

PROMPT = "我正在处理我的课程表信息，现在需要你来协助我将我的课表信息转换为我的课表软件能够识别的格式，我将给你提供课程字段，你需要将这些字段转换为‘课程名称,星期,开始节数,结束节数,老师,地点,周数’的格式。注意，每个字段可能有多个课程，这种情况你需要输出两条数据，用回车隔开；同时，有些数据比如老师可能有缺失，这种情况输出空格。你的回答不需要包含除了输出格式以外的任何信息，以下是两个例子：我提供的字段为‘星期一 信息论\n[电信待定1]\n[9-10,12-14,16周][H302]\n第1-2节’，你需要输出‘信息论,1,1,2, ,H302,9-10、12-14、16’。假设我提供的字段是'星期二 【实验】信号与系统 实验分组1-4/第6组\n[1-4节][14-14周]\n[K405]\n计算机视觉\n[苏敬勇]\n[5-13周][H411]\n第1-2节'，你需要输出'【实验】信号与系统,2,1,4, ,K405,14\n计算机视觉,2,1,2,苏敬勇,H411,5-13’。现在，我给你的字段是："

def process_schedule(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=4, min_col=2, max_col=sheet.max_column, max_row=sheet.max_row):
        for cell in row:
            if cell.value:
                day_of_week = {
                    2: "星期一",
                    3: "星期二",
                    4: "星期三",
                    5: "星期四",
                    6: "星期五",
                    7: "星期六",
                    8: "星期日"
                }
                day = day_of_week.get(cell.column, None)
                if day:
                    data.append(day + ' ' + cell.value)
    return data

client = openai.OpenAI(
    base_url=BASE_URL,
    api_key=API_KEY
)

table = process_schedule('export.xlsx')

with open('class.csv', 'w', encoding='utf-8') as f:
    f.write('课程名称,星期,开始节数,结束节数,老师,地点,周数\n')

for data in tqdm(table):
    completion = client.chat.completions.create(
    model="gpt-4o",
    messages=[
        {"role": "system", "content": "你将协助用户转换字段数据。"},
        {"role": "user", "content": PROMPT + data}
    ]
    )
    with open('class.csv', 'a', encoding='utf-8') as f:
        f.write(completion.choices[0].message.content + '\n')

print('=====Finished!=====')